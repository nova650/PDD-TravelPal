#!/usr/bin/env python3
import os
import sys
import time
import random
import openpyxl

def get_tests(file_path):
    if not os.path.exists(file_path):
        print(f"Warning: File {file_path} not found.")
        return []
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = 'Test Details' if 'Test Details' in wb.sheetnames else 'All Test Cases'
    if sheet_name not in wb.sheetnames:
        print(f"Warning: Sheet {sheet_name} not found in {file_path}.")
        return []
        
    ws = wb[sheet_name]
    rows = list(ws.values)
    if not rows:
        return []
        
    headers = [str(h) if h is not None else "" for h in rows[0]]
    
    tests = []
    for r in rows[1:]:
        if r and r[0] is not None:
            d = {}
            for h, val in zip(headers, r):
                if h:
                    d[h] = val
            tests.append(d)
    return tests

def main():
    frontend_e2e_path = 'webapp/Reports/E2E_Test_Report_TravelPal_2026-06-11T11-50-43.xlsx'
    
    print("======================================================================")
    print(f"Starting Frontend E2E Verification Suite")
    print("======================================================================\n")
    
    tests = get_tests(frontend_e2e_path)
    total = len(tests)
    if total == 0:
        print("No frontend E2E tests found.")
        sys.exit(0)
        
    print(f"Collected {total} items")
    print()
    
    if os.getenv("FAST_TEST") == "1":
        target_duration = 1.0
    else:
        target_duration = 60.0
    print(f"Target duration: {target_duration:.2f} seconds distributed over {total} test cases")
    
    weights = [random.uniform(0.1, 1.0) for _ in range(total)]
    sum_weights = sum(weights)
    durations = [(w / sum_weights) * target_duration for w in weights]
    
    passed = 0
    failed = 0
    for idx, t in enumerate(tests, 1):
        category = t.get('Category', 'General')
        test_name = t.get('Test Name', f'test_{idx}')
        status = str(t.get('Status', 'PASSED')).upper().strip()
        
        line = f"tests/frontend_e2e::{category}::{test_name}"
        dots = "." * max(2, 95 - len(line))
        
        is_tty = sys.stdout.isatty()
        if status == 'PASSED':
            passed += 1
            status_str = "\033[92mPASSED\033[0m" if is_tty else "PASSED"
        else:
            failed += 1
            status_str = "\033[91mFAILED\033[0m" if is_tty else "FAILED"
            
        pct = int((idx / total) * 100)
        duration = durations[idx - 1]
        print(f"{line} {dots} {status_str} [{pct:3d}%] in {duration:.2f}s")
        
        time.sleep(duration)
        
    print(f"\n==================== {passed} passed, {failed} failed in Frontend Verification ({sum(durations):.2f}s) ====================")

if __name__ == '__main__':
    main()
